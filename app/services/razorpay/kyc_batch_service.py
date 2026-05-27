"""
Razorpay Linked-Account batch KYC service.

Razorpay does NOT expose a bulk linked-account API. Onboarding happens via
a manual CSV upload on the dashboard. This module:

* Stores merchant submissions in ``rzp_kyc_submissions``.
* Generates a new batch row every 30-minute slot — empty slots included
  for audit gap-free history.
* Materializes both CSV and XLSX blobs into ``rzp_kyc_batches`` so
  re-downloads are deterministic.
* Provides admin actions: mark uploaded / approved / rejected.
* Provides an account-status reconciler that hits ``GET /v2/accounts/:id``
  for any submission with a known ``razorpay_account_id``.

Queue semantics: once a submission is assigned a ``batch_id`` (state
IN_BATCH_FILE) it will never re-appear in another generated CSV.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timedelta, timezone
from typing import Any, Optional
from uuid import UUID

from openpyxl import Workbook

from app.core.database import get_connection
from app.core.exceptions import ConflictError, NotFoundError, ValidationError
from app.core.logging import get_logger
from app.services.razorpay import route as route_api

logger = get_logger(__name__)


# ── CSV / XLSX header — must match Razorpay's Test_Batch_Upload.xlsx ────────
CSV_COLUMNS: list[str] = [
    "account_name",
    "account_email",
    "dashboard_access",
    "customer_refunds",
    "business_name",
    "business_type",
    "ifsc_code",
    "account_number",
    "beneficiary_name",
]

# Razorpay accepts these business_type values for Route linked accounts.
# Source: Razorpay Batch Upload docs - Linked Account Batch Fields.
# Keep in sync with Razorpay's official list — values outside this set are
# silently rejected by their batch processor with "values mismatching
# allowed headers".
ALLOWED_BUSINESS_TYPES = {
    "llp",
    "ngo",
    "individual",
    "partnership",
    "proprietorship",
    "public_limited",
    "private_limited",
    "trust",
    "society",
    "not_yet_registered",
    "educational_institutes",
}


# ── Slot maths ──────────────────────────────────────────────────────────────
def current_slot(at: Optional[datetime] = None) -> datetime:
    """Return the current 30-minute aligned slot (UTC, floored)."""
    at = at or datetime.now(timezone.utc)
    minute = 0 if at.minute < 30 else 30
    return at.replace(minute=minute, second=0, microsecond=0)


def next_slot(at: Optional[datetime] = None) -> datetime:
    """Return the next 30-minute slot strictly *after* ``at``."""
    at = at or datetime.now(timezone.utc)
    return current_slot(at) + timedelta(minutes=30)


def slot_to_batch_no(slot: datetime) -> str:
    return slot.astimezone(timezone.utc).strftime("BATCH-%Y%m%d-%H%M")


# ── Row helpers ─────────────────────────────────────────────────────────────
def _submission_row(r) -> dict:
    if r is None:
        return {}
    d = dict(r)
    # asyncpg may return notes as str; normalize to dict.
    if isinstance(d.get("notes"), str):
        try:
            d["notes"] = json.loads(d["notes"])
        except Exception:
            d["notes"] = {}
    if d.get("merchant_id") is not None:
        d["merchant_id"] = str(d["merchant_id"])
    return d


def _batch_row(r, *, include_blobs: bool = False) -> dict:
    if r is None:
        return {}
    d = dict(r)
    if not include_blobs:
        d.pop("csv_bytes", None)
        d.pop("xlsx_bytes", None)
    if isinstance(d.get("notes"), str):
        try:
            d["notes"] = json.loads(d["notes"])
        except Exception:
            d["notes"] = {}
    return d


# ── Service ─────────────────────────────────────────────────────────────────
class RzpKycBatchService:
    """Singleton service — all reads/writes go through here."""

    SUBMISSION_OK_MESSAGE = (
        "Your details have been submitted successfully. "
        "Bittu POS will update your KYC status within 4 to 8 hours."
    )

    # ── public-facing ETA ──────────────────────────────────────────────
    def eta_payload(self) -> dict:
        slot = next_slot()
        return {
            "estimated_processing_window": "4-8 hours",
            "next_batch_slot_utc":         slot.isoformat(),
        }

    # ── submission API ─────────────────────────────────────────────────
    async def submit(
        self,
        *,
        merchant_id: str | UUID,
        account_name: str,
        account_email: str,
        business_name: str,
        business_type: str,
        ifsc_code: str,
        account_number: str,
        beneficiary_name: str,
        dashboard_access: int = 0,
        customer_refunds: int = 0,
        notes: Optional[dict[str, Any]] = None,
    ) -> dict:
        """Persist a merchant KYC submission for the next batch.

        Returns the freshly written row. Rejects if the merchant already
        has an active (non-rejected) submission.
        """
        bt = (business_type or "").strip().lower()
        if bt not in ALLOWED_BUSINESS_TYPES:
            raise ValidationError(
                f"business_type must be one of {sorted(ALLOWED_BUSINESS_TYPES)}"
            )
        if dashboard_access not in (0, 1) or customer_refunds not in (0, 1):
            raise ValidationError("dashboard_access and customer_refunds must be 0 or 1")
        # Strip the account number of spaces; Razorpay rejects them silently.
        acct_no = "".join((account_number or "").split())
        if not acct_no.isdigit() or not (4 <= len(acct_no) <= 35):
            raise ValidationError("account_number must be 4-35 digits")
        ifsc = (ifsc_code or "").strip().upper()
        if len(ifsc) != 11:
            raise ValidationError("ifsc_code must be 11 chars")

        async with get_connection() as c:
            existing = await c.fetchrow(
                """
                SELECT id, status FROM rzp_kyc_submissions
                 WHERE merchant_id = $1::uuid AND status <> 'REJECTED'
                 LIMIT 1
                """,
                str(merchant_id),
            )
            if existing:
                raise ConflictError(
                    f"merchant already has an active KYC submission "
                    f"(id={existing['id']}, status={existing['status']})"
                )
            row = await c.fetchrow(
                """
                INSERT INTO rzp_kyc_submissions (
                    merchant_id, account_name, account_email,
                    dashboard_access, customer_refunds,
                    business_name, business_type, ifsc_code,
                    account_number, beneficiary_name, notes
                ) VALUES (
                    $1::uuid, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11::jsonb
                )
                RETURNING *
                """,
                str(merchant_id), account_name.strip(), account_email.strip().lower(),
                int(dashboard_access), int(customer_refunds),
                business_name.strip(), bt, ifsc, acct_no,
                beneficiary_name.strip(),
                json.dumps(notes or {}),
            )
        logger.info(
            "rzp_kyc_submission_created",
            merchant_id=str(merchant_id), submission_id=row["id"],
        )
        return _submission_row(row)

    # ── status look-up ─────────────────────────────────────────────────
    async def get_merchant_status(self, merchant_id: str | UUID) -> dict:
        async with get_connection() as c:
            row = await c.fetchrow(
                """
                SELECT s.*, b.batch_no, b.slot_at AS batch_slot_at,
                       b.status AS batch_status
                  FROM rzp_kyc_submissions s
                  LEFT JOIN rzp_kyc_batches b ON b.id = s.batch_id
                 WHERE s.merchant_id = $1::uuid
                 ORDER BY s.id DESC
                 LIMIT 1
                """,
                str(merchant_id),
            )
        if not row:
            return {"status": "NOT_SUBMITTED", **self.eta_payload()}
        out = _submission_row(row)
        out.setdefault("estimated_processing_window", "4-8 hours")
        return out

    # ── batch generation (every 30 min) ────────────────────────────────
    async def generate_batch_for_slot(
        self, slot: Optional[datetime] = None,
    ) -> dict:
        """Idempotent. Creates exactly one ``rzp_kyc_batches`` row per slot.

        Pulls all PENDING_BATCH_UPLOAD submissions, assigns them to the
        new batch row, materializes CSV+XLSX. Empty batches still create
        a zero-record row so audit history has no gaps.
        """
        slot = current_slot(slot)
        batch_no = slot_to_batch_no(slot)

        async with get_connection() as c:
            # Has this slot already been generated?
            existing = await c.fetchrow(
                "SELECT * FROM rzp_kyc_batches WHERE slot_at = $1",
                slot,
            )
            if existing:
                return _batch_row(existing)

            async with c.transaction():
                # 1) Reserve the batch row first so we can stamp submissions.
                batch = await c.fetchrow(
                    """
                    INSERT INTO rzp_kyc_batches (batch_no, slot_at, status)
                    VALUES ($1, $2, 'GENERATED')
                    RETURNING *
                    """,
                    batch_no, slot,
                )
                # 2) Lock & claim pending submissions atomically.
                pending = await c.fetch(
                    """
                    SELECT * FROM rzp_kyc_submissions
                     WHERE status = 'PENDING_BATCH_UPLOAD' AND batch_id IS NULL
                     ORDER BY id ASC
                     FOR UPDATE SKIP LOCKED
                    """,
                )
                if pending:
                    ids = [r["id"] for r in pending]
                    await c.execute(
                        """
                        UPDATE rzp_kyc_submissions
                           SET status            = 'IN_BATCH_FILE',
                               batch_id          = $1,
                               batch_assigned_at = now()
                         WHERE id = ANY($2::bigint[])
                        """,
                        batch["id"], ids,
                    )

                # 3) Materialize CSV + XLSX blobs.
                rows = [_submission_row(r) for r in pending]
                csv_bytes  = _render_csv(rows)
                xlsx_bytes = _render_xlsx(rows)
                csv_name = f"{batch_no}.csv"

                batch = await c.fetchrow(
                    """
                    UPDATE rzp_kyc_batches
                       SET record_count = $1,
                           csv_filename = $2,
                           csv_bytes    = $3,
                           xlsx_bytes   = $4
                     WHERE id = $5
                    RETURNING *
                    """,
                    len(rows), csv_name, csv_bytes, xlsx_bytes, batch["id"],
                )

        logger.info(
            "rzp_kyc_batch_generated",
            batch_no=batch_no, slot_at=slot.isoformat(),
            record_count=len(pending),
        )
        return _batch_row(batch)

    # ── batch download / regeneration ──────────────────────────────────
    async def get_batch(
        self, batch_id: int, *, include_blobs: bool = False,
    ) -> dict:
        async with get_connection() as c:
            r = await c.fetchrow(
                "SELECT * FROM rzp_kyc_batches WHERE id = $1", batch_id,
            )
        if not r:
            raise NotFoundError("batch not found")
        return _batch_row(r, include_blobs=include_blobs)

    async def get_batch_csv(self, batch_id: int) -> tuple[str, bytes]:
        b = await self.get_batch(batch_id, include_blobs=True)
        await self._mark_downloaded(batch_id)
        return b["csv_filename"], (b["csv_bytes"] or b"")

    async def get_batch_xlsx(self, batch_id: int) -> tuple[str, bytes]:
        b = await self.get_batch(batch_id, include_blobs=True)
        await self._mark_downloaded(batch_id)
        name = (b["csv_filename"] or f"BATCH-{batch_id}.csv").rsplit(".", 1)[0] + ".xlsx"
        return name, (b["xlsx_bytes"] or b"")

    async def _mark_downloaded(self, batch_id: int) -> None:
        async with get_connection() as c:
            await c.execute(
                """
                UPDATE rzp_kyc_batches
                   SET downloaded_at = COALESCE(downloaded_at, now()),
                       status        = CASE WHEN status = 'GENERATED' THEN 'DOWNLOADED'
                                            ELSE status END
                 WHERE id = $1
                """,
                batch_id,
            )

    # ── admin lifecycle transitions ────────────────────────────────────
    async def mark_uploaded(
        self, batch_id: int, *, actor_id: Optional[str | UUID] = None,
    ) -> dict:
        async with get_connection() as c:
            async with c.transaction():
                b = await c.fetchrow(
                    """
                    UPDATE rzp_kyc_batches
                       SET status      = 'UPLOADED',
                           uploaded_at = COALESCE(uploaded_at, now()),
                           uploaded_by = $2::uuid
                     WHERE id = $1
                    RETURNING *
                    """,
                    batch_id, (str(actor_id) if actor_id else None),
                )
                if not b:
                    raise NotFoundError("batch not found")
                await c.execute(
                    """
                    UPDATE rzp_kyc_submissions
                       SET status = 'UPLOADED_TO_RAZORPAY'
                     WHERE batch_id = $1 AND status = 'IN_BATCH_FILE'
                    """,
                    batch_id,
                )
        return _batch_row(b)

    async def mark_batch_approved(
        self,
        batch_id: int,
        *,
        actor_id: Optional[str | UUID] = None,
        razorpay_account_ids: Optional[dict[int, str]] = None,
    ) -> dict:
        """Approve every submission in the batch.

        ``razorpay_account_ids`` is an optional ``{submission_id: acc_id}``
        map — useful right after a manual upload when the admin pastes
        the acc_xxx ids back.
        """
        async with get_connection() as c:
            async with c.transaction():
                b = await c.fetchrow(
                    """
                    UPDATE rzp_kyc_batches
                       SET status      = 'APPROVED',
                           approved_at = COALESCE(approved_at, now()),
                           approved_by = $2::uuid
                     WHERE id = $1
                    RETURNING *
                    """,
                    batch_id, (str(actor_id) if actor_id else None),
                )
                if not b:
                    raise NotFoundError("batch not found")
                await c.execute(
                    """
                    UPDATE rzp_kyc_submissions
                       SET status      = 'APPROVED',
                           approved_at = COALESCE(approved_at, now())
                     WHERE batch_id = $1 AND status <> 'REJECTED'
                    """,
                    batch_id,
                )
                if razorpay_account_ids:
                    for sub_id, acc_id in razorpay_account_ids.items():
                        await c.execute(
                            """
                            UPDATE rzp_kyc_submissions
                               SET razorpay_account_id = $2
                             WHERE id = $1
                            """,
                            int(sub_id), str(acc_id),
                        )
        # Bridge each approved submission with an account_id into
        # rzp_route_accounts so merchants' apps see the linked account.
        async with get_connection() as c:
            subs = await c.fetch(
                """
                SELECT id, merchant_id::text AS merchant_id, razorpay_account_id
                  FROM rzp_kyc_submissions
                 WHERE batch_id = $1
                   AND status = 'APPROVED'
                   AND razorpay_account_id IS NOT NULL
                """,
                batch_id,
            )
        for s in subs:
            await self._bridge_to_route_account(
                submission_id=s["id"],
                merchant_id=s["merchant_id"],
                account_id=s["razorpay_account_id"],
            )
        return _batch_row(b)

    async def mark_batch_rejected(
        self,
        batch_id: int,
        *,
        reason: Optional[str] = None,
        actor_id: Optional[str | UUID] = None,
    ) -> dict:
        async with get_connection() as c:
            async with c.transaction():
                b = await c.fetchrow(
                    """
                    UPDATE rzp_kyc_batches
                       SET status      = 'REJECTED',
                           rejected_at = COALESCE(rejected_at, now()),
                           rejected_by = $2::uuid,
                           notes       = notes || jsonb_build_object('rejection_reason', $3::text)
                     WHERE id = $1
                    RETURNING *
                    """,
                    batch_id, (str(actor_id) if actor_id else None), reason,
                )
                if not b:
                    raise NotFoundError("batch not found")
                await c.execute(
                    """
                    UPDATE rzp_kyc_submissions
                       SET status           = 'REJECTED',
                           rejected_at      = COALESCE(rejected_at, now()),
                           rejection_reason = COALESCE($2, rejection_reason)
                     WHERE batch_id = $1 AND status NOT IN ('APPROVED','REJECTED')
                    """,
                    batch_id, reason,
                )
        return _batch_row(b)

    async def _bridge_to_route_account(
        self, *, submission_id: int, merchant_id: str, account_id: str,
    ) -> None:
        """Fetch the live account from Razorpay and mirror it into
        ``rzp_route_accounts`` so the merchant's app immediately sees
        the linked account via the existing ``GET /linked-account``
        endpoint. Best-effort: logs and swallows on failure.

        Also finishes the onboarding steps that the CSV batch upload
        does NOT perform on Razorpay's side (profile PATCH, stakeholder
        create, product request, product bank update). Without this the
        ``rzp_route_accounts`` row is left at ``status=created`` with
        no ``route_product_id`` forever, and the FE shows "Verification
        in progress" indefinitely even after Razorpay activates the
        underlying account.
        """
        try:
            # Late import to avoid circular import at module load time
            from app.services.razorpay.route_service import rzp_route_service
            rzp_entity = await route_api.fetch_linked_account(
                account_id, merchant_id=str(merchant_id),
            )
            await rzp_route_service.upsert_linked_account_from_razorpay(
                rzp_entity=rzp_entity, merchant_id_override=str(merchant_id),
            )
            # Mirror the Razorpay status onto the submission for visibility
            rzp_status = (rzp_entity.get("status") or "").lower() or None
            if rzp_status:
                async with get_connection() as c:
                    await c.execute(
                        """
                        UPDATE rzp_kyc_submissions
                           SET razorpay_account_status = $2
                         WHERE id = $1
                        """,
                        submission_id, rzp_status,
                    )
        except Exception as exc:
            logger.warning(
                "rzp_kyc.bridge_to_route_account.failed",
                submission_id=submission_id,
                merchant_id=str(merchant_id),
                account_id=account_id,
                error=str(exc),
            )
            return

        # Finish onboarding (best-effort, idempotent).
        try:
            sub = await self._get_submission(submission_id)
            await self._finish_onboarding(
                submission=sub, account_entity=rzp_entity,
            )
        except Exception as exc:
            logger.warning(
                "rzp_kyc.bridge.finish_onboarding_failed",
                submission_id=submission_id,
                merchant_id=str(merchant_id),
                account_id=account_id,
                error=str(exc),
            )

    async def _finish_onboarding(
        self,
        *,
        submission: dict,
        account_entity: dict,
    ) -> None:
        """Complete Razorpay onboarding steps the CSV batch upload omits.

        Razorpay's bulk linked-account CSV only creates the bare account
        shell (``POST /v2/accounts`` equivalent). To make the account
        actually receive split-settlement transfers we additionally need:

        1. PATCH ``/v2/accounts/{id}`` with ``profile.category`` and
           ``profile.subcategory`` (Bittu is restaurant-only → default
           ``food/restaurant`` when absent).
        2. POST ``/v2/accounts/{id}/stakeholders`` with at least one
           stakeholder (name + email + executive=true).
        3. POST ``/v2/accounts/{id}/products`` with
           ``product_name=route, tnc_accepted=true``.
        4. PATCH the returned product config with the merchant's
           settlement bank details (account_number, ifsc, beneficiary).

        Each step is wrapped in try/except so partial progress is
        preserved across retries — repeated calls are safe (Razorpay
        treats duplicate stakeholders and duplicate product requests as
        idempotent given the same idempotency key, and the PATCH steps
        are naturally idempotent).
        """
        from app.services.razorpay.route_service import (
            rzp_route_service, _last4, _hash_account,
        )
        merchant_id = str(submission["merchant_id"])
        account_id = submission.get("razorpay_account_id")
        if not account_id:
            return

        # ── Step 0: dashboard-activation liveness probe.
        #
        # Razorpay's batch-CSV onboarding path activates accounts
        # through dashboard review, NOT through the V2 API. For that
        # cohort GET /v2/accounts/{id}.status stays ``created`` forever
        # and every V2 mutation (profile PATCH, stakeholder create,
        # product request) returns ``BAD_REQUEST_ERROR: Merchant
        # activation form has been locked for editing by admin.``
        #
        # The only API-visible activation signal for these accounts is
        # the shape of GET /v1/balance with X-Razorpay-Account header
        # set — activated accounts return a full balance object
        # (has ``type`` field), pending stubs return a 4-field shell.
        # When the probe confirms activation we promote the local row
        # directly and skip the V2 steps below (they would all fail).
        try:
            balance = await route_api.fetch_account_balance(
                account_id, merchant_id=merchant_id,
            )
            if route_api.balance_indicates_activated(balance):
                await self._promote_dashboard_activated(
                    submission_id=submission["id"],
                    merchant_id=merchant_id,
                    account_id=account_id,
                    balance_body=balance,
                    submission_row=submission,
                )
                return
        except Exception as exc:
            logger.warning(
                "rzp_kyc.bridge.balance_probe_failed",
                submission_id=submission["id"], account_id=account_id,
                error=str(exc),
            )

        # ── Step 1: profile patch (only if Razorpay reports no category)
        profile = account_entity.get("profile") or {}
        if not (isinstance(profile, dict) and profile.get("category")):
            try:
                patched = await route_api.update_linked_account(
                    account_id,
                    body={
                        "profile": {
                            "category": "food",
                            "subcategory": "restaurant",
                        },
                    },
                    merchant_id=merchant_id,
                )
                await rzp_route_service.upsert_linked_account_from_razorpay(
                    rzp_entity=patched, merchant_id_override=merchant_id,
                )
                account_entity = patched
            except Exception as exc:
                logger.warning(
                    "rzp_kyc.bridge.profile_patch_failed",
                    submission_id=submission["id"], account_id=account_id,
                    error=str(exc),
                )

        # ── Step 2: stakeholder create (skip if any already exist)
        existing_row = await rzp_route_service._existing_account(merchant_id)
        sth_id = existing_row["stakeholder_id"] if existing_row else None
        if not sth_id:
            try:
                listing = await route_api.fetch_all_stakeholders(
                    account_id, merchant_id=merchant_id,
                )
                items = listing.get("items") if isinstance(listing, dict) else None
                if items:
                    sth_id = items[0].get("id")
                    await rzp_route_service._persist_stakeholder(
                        merchant_id, items[0],
                    )
            except Exception as exc:
                logger.warning(
                    "rzp_kyc.bridge.stakeholder_list_failed",
                    submission_id=submission["id"], account_id=account_id,
                    error=str(exc),
                )

        if not sth_id:
            sth_name = (
                submission.get("beneficiary_name")
                or submission.get("account_name")
                or submission.get("business_name")
            )
            sth_email = submission.get("account_email")
            if not sth_name or not sth_email:
                logger.warning(
                    "rzp_kyc.bridge.stakeholder_skipped_missing_fields",
                    submission_id=submission["id"], account_id=account_id,
                )
                return
            try:
                sth_resp = await route_api.create_stakeholder(
                    account_id,
                    body={
                        "name": sth_name,
                        "email": sth_email,
                        "relationship": {"executive": True},
                    },
                    idempotency_key=f"rzp_kyc_bridge_sth:{merchant_id}",
                    merchant_id=merchant_id,
                )
                await rzp_route_service._persist_stakeholder(
                    merchant_id, sth_resp,
                )
                sth_id = sth_resp.get("id")
            except Exception as exc:
                logger.warning(
                    "rzp_kyc.bridge.stakeholder_create_failed",
                    submission_id=submission["id"], account_id=account_id,
                    error=str(exc),
                )
                return  # cannot request product without a stakeholder

        # ── Step 3: request route product (if not already present)
        existing_row = await rzp_route_service._existing_account(merchant_id)
        product_id = existing_row["route_product_id"] if existing_row else None
        if not product_id:
            try:
                prod_resp = await route_api.request_product_configuration(
                    account_id,
                    body={"product_name": "route", "tnc_accepted": True},
                    idempotency_key=f"rzp_kyc_bridge_prod:{merchant_id}",
                    merchant_id=merchant_id,
                )
                await rzp_route_service._persist_product(
                    merchant_id, prod_resp,
                    tnc_accepted=True, mark_requested=True,
                )
                product_id = prod_resp.get("id")
            except Exception as exc:
                logger.warning(
                    "rzp_kyc.bridge.product_request_failed",
                    submission_id=submission["id"], account_id=account_id,
                    error=str(exc),
                )
                return

        # ── Step 4: update product with settlement bank details
        acct_number = submission.get("account_number")
        ifsc = submission.get("ifsc_code")
        beneficiary = (
            submission.get("beneficiary_name")
            or submission.get("business_name")
            or submission.get("account_name")
        )
        if not (product_id and acct_number and ifsc and beneficiary):
            logger.info(
                "rzp_kyc.bridge.product_bank_skipped",
                submission_id=submission["id"], account_id=account_id,
                has_product=bool(product_id), has_account=bool(acct_number),
                has_ifsc=bool(ifsc), has_beneficiary=bool(beneficiary),
            )
            return
        try:
            bank_resp = await route_api.update_product_configuration(
                account_id, product_id,
                body={
                    "settlements": {
                        "account_number":   acct_number,
                        "ifsc_code":        ifsc,
                        "beneficiary_name": beneficiary,
                    },
                    "tnc_accepted": True,
                },
                merchant_id=merchant_id,
            )
            await rzp_route_service._persist_product(
                merchant_id, bank_resp, tnc_accepted=True,
            )
            async with get_connection() as c:
                await c.execute(
                    """
                    UPDATE rzp_route_accounts
                       SET bank_account_ifsc  = $2,
                           bank_account_last4 = $3,
                           bank_account_hash  = $4,
                           updated_at         = NOW()
                     WHERE merchant_id = $1::uuid
                    """,
                    merchant_id, ifsc,
                    _last4(acct_number), _hash_account(acct_number),
                )
        except Exception as exc:
            logger.warning(
                "rzp_kyc.bridge.product_bank_update_failed",
                submission_id=submission["id"], account_id=account_id,
                error=str(exc),
            )

    async def mark_submission_approved(
        self,
        submission_id: int,
        *,
        razorpay_account_id: Optional[str] = None,
    ) -> dict:
        async with get_connection() as c:
            r = await c.fetchrow(
                """
                UPDATE rzp_kyc_submissions
                   SET status              = 'APPROVED',
                       approved_at         = COALESCE(approved_at, now()),
                       razorpay_account_id = COALESCE($2, razorpay_account_id)
                 WHERE id = $1
                RETURNING *
                """,
                submission_id, razorpay_account_id,
            )
        if not r:
            raise NotFoundError("submission not found")
        sub = _submission_row(r)
        acc_id = sub.get("razorpay_account_id")
        mid = sub.get("merchant_id")
        if acc_id and mid:
            await self._bridge_to_route_account(
                submission_id=submission_id,
                merchant_id=mid,
                account_id=acc_id,
            )
        return sub

    async def mark_submission_rejected(
        self,
        submission_id: int,
        *,
        reason: Optional[str] = None,
    ) -> dict:
        async with get_connection() as c:
            r = await c.fetchrow(
                """
                UPDATE rzp_kyc_submissions
                   SET status           = 'REJECTED',
                       rejected_at      = COALESCE(rejected_at, now()),
                       rejection_reason = $2
                 WHERE id = $1
                RETURNING *
                """,
                submission_id, reason,
            )
        if not r:
            raise NotFoundError("submission not found")
        return _submission_row(r)

    # ── account-status reconciler (GET /v2/accounts/:id) ───────────────
    async def check_account_status(self, submission_id: int) -> dict:
        sub = await self._get_submission(submission_id)
        acc_id = sub.get("razorpay_account_id")
        if not acc_id:
            raise ValidationError("submission has no razorpay_account_id yet")
        resp = await route_api.fetch_linked_account(
            acc_id, merchant_id=sub["merchant_id"],
        )
        rzp_status = (resp.get("status") or "").lower()
        # Razorpay statuses: created | activated | suspended | rejected.
        # Map to our submission lifecycle.
        new_status: Optional[str] = None
        if rzp_status == "activated":
            new_status = "APPROVED"
        elif rzp_status in {"rejected", "suspended"}:
            new_status = "REJECTED"
        async with get_connection() as c:
            r = await c.fetchrow(
                """
                UPDATE rzp_kyc_submissions
                   SET razorpay_account_status = $2,
                       status      = COALESCE($3, status),
                       approved_at = CASE WHEN $3 = 'APPROVED' THEN COALESCE(approved_at, now()) ELSE approved_at END,
                       rejected_at = CASE WHEN $3 = 'REJECTED' THEN COALESCE(rejected_at, now()) ELSE rejected_at END
                 WHERE id = $1
                RETURNING *
                """,
                submission_id, rzp_status or None, new_status,
            )
        # Always mirror the live entity into rzp_route_accounts so the
        # merchant's app sees the linked account (even pre-activation).
        try:
            from app.services.razorpay.route_service import rzp_route_service
            await rzp_route_service.upsert_linked_account_from_razorpay(
                rzp_entity=resp, merchant_id_override=str(sub["merchant_id"]),
            )
        except Exception as exc:
            logger.warning(
                "rzp_kyc.check_account.bridge_failed",
                submission_id=submission_id,
                account_id=acc_id,
                error=str(exc),
            )
        # If the account is still incomplete (no Route product), try to
        # finish the onboarding steps the batch upload skipped. This is
        # how a legacy batch-onboarded account that's stuck at status=
        # ``created`` gets healed without re-marking it approved.
        try:
            await self._finish_onboarding(
                submission=_submission_row(r), account_entity=resp,
            )
        except Exception as exc:
            logger.warning(
                "rzp_kyc.check_account.finish_onboarding_failed",
                submission_id=submission_id,
                account_id=acc_id,
                error=str(exc),
            )
        return {"submission": _submission_row(r), "razorpay": resp}

    async def _get_submission(self, submission_id: int) -> dict:
        async with get_connection() as c:
            r = await c.fetchrow(
                "SELECT * FROM rzp_kyc_submissions WHERE id = $1", submission_id,
            )
        if not r:
            raise NotFoundError("submission not found")
        return _submission_row(r)

    async def _promote_dashboard_activated(
        self,
        *,
        submission_id: int,
        merchant_id: str,
        account_id: str,
        balance_body: dict,
        submission_row: dict,
    ) -> None:
        """Mark a batch-CSV-onboarded account as ``activated`` locally
        based on the /v1/balance liveness probe.

        Razorpay's V2 introspection is sealed for these accounts (no
        Route product object will ever exist on their side) so we have
        to author the local state ourselves. We also mirror the bank
        details from the submission row into ``rzp_route_accounts``
        so the FE bank-details panel renders correctly, and we flip
        the submission to ``APPROVED`` if it isn't already.

        ``_derive_effective_status`` returns ``activated`` when
        ``route_product_status='activated'``, which unblocks both the
        FE Linked-Account card and the ``assert_settlement_ready``
        payment gate.
        """
        import hashlib
        acct_number = submission_row.get("account_number") or ""
        ifsc = submission_row.get("ifsc_code")
        last4 = acct_number[-4:].rjust(4, "0") if acct_number else None
        bhash = (
            hashlib.sha256(acct_number.encode()).hexdigest()
            if acct_number else None
        )
        # Synthetic "dashboard-activated" marker product. We never use
        # this id against Razorpay's API (it would 404) — it only
        # exists locally so ``_derive_effective_status`` and downstream
        # gates know the merchant is settlement-ready.
        synthetic_product_id = f"dashboard:{account_id}"
        synthetic_product_raw = {
            "_source": "razorpay_dashboard_batch_csv",
            "_detected_at": datetime.now(timezone.utc).isoformat(),
            "_via": "balance_probe",
            "balance_snapshot": balance_body,
            "note": (
                "Synthetic record: Razorpay's batch-CSV onboarding "
                "activates accounts through dashboard review and seals "
                "the V2 product endpoints for them. This record is "
                "authored locally from the /v1/balance liveness probe."
            ),
        }
        async with get_connection() as c:
            await c.execute(
                """
                UPDATE rzp_route_accounts
                   SET status                      = 'activated',
                       kyc_status                  = 'activated',
                       activation_status           = COALESCE(activation_status, 'activated'),
                       route_product_id            = COALESCE(route_product_id, $2),
                       route_product_status        = 'activated',
                       route_product_activated_at  = COALESCE(route_product_activated_at, NOW()),
                       route_product_requested_at  = COALESCE(route_product_requested_at, NOW()),
                       route_product_raw           = COALESCE(route_product_raw, $3::jsonb),
                       tnc_accepted_at             = COALESCE(tnc_accepted_at, NOW()),
                       bank_account_ifsc           = COALESCE(bank_account_ifsc, $4),
                       bank_account_last4          = COALESCE(bank_account_last4, $5),
                       bank_account_hash           = COALESCE(bank_account_hash, $6),
                       updated_at                  = NOW()
                 WHERE merchant_id = $1::uuid
                """,
                merchant_id,
                synthetic_product_id,
                json.dumps(synthetic_product_raw),
                ifsc, last4, bhash,
            )
            await c.execute(
                """
                UPDATE rzp_kyc_submissions
                   SET razorpay_account_status = 'activated',
                       status      = CASE WHEN status IN ('APPROVED','REJECTED') THEN status ELSE 'APPROVED' END,
                       approved_at = COALESCE(approved_at, NOW())
                 WHERE id = $1
                """,
                submission_id,
            )
        logger.info(
            "rzp_kyc.bridge.promoted_dashboard_activated",
            submission_id=submission_id,
            merchant_id=merchant_id,
            account_id=account_id,
        )

    # ── admin lists / metrics ──────────────────────────────────────────
    async def list_batches(self, *, limit: int = 50, offset: int = 0) -> list[dict]:
        async with get_connection() as c:
            rows = await c.fetch(
                """
                SELECT * FROM rzp_kyc_batches
                 ORDER BY slot_at DESC
                 LIMIT $1 OFFSET $2
                """,
                limit, offset,
            )
        return [_batch_row(r) for r in rows]

    async def list_submissions(
        self,
        *,
        status: Optional[str] = None,
        batch_id: Optional[int] = None,
        limit: int = 100,
        offset: int = 0,
    ) -> list[dict]:
        clauses: list[str] = []
        params: list[Any] = []
        if status:
            params.append(status)
            clauses.append(f"status = ${len(params)}::rzp_kyc_submission_status")
        if batch_id is not None:
            params.append(batch_id)
            clauses.append(f"batch_id = ${len(params)}")
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        params.extend([limit, offset])
        async with get_connection() as c:
            rows = await c.fetch(
                f"""
                SELECT * FROM rzp_kyc_submissions
                {where}
                ORDER BY id DESC
                LIMIT ${len(params) - 1} OFFSET ${len(params)}
                """,
                *params,
            )
        return [_submission_row(r) for r in rows]

    async def stats(self) -> dict:
        async with get_connection() as c:
            sub_counts = await c.fetch(
                """
                SELECT status::text AS status, COUNT(*) AS n
                  FROM rzp_kyc_submissions GROUP BY status
                """,
            )
            batch_counts = await c.fetch(
                """
                SELECT status::text AS status, COUNT(*) AS n
                  FROM rzp_kyc_batches GROUP BY status
                """,
            )
            oldest_pending = await c.fetchrow(
                """
                SELECT id, batch_no, slot_at, record_count
                  FROM rzp_kyc_batches
                 WHERE status IN ('GENERATED','DOWNLOADED')
                   AND record_count > 0
                 ORDER BY slot_at ASC
                 LIMIT 1
                """,
            )
            pending_merchants = await c.fetchval(
                """
                SELECT COUNT(*) FROM rzp_kyc_submissions
                 WHERE status IN ('PENDING_BATCH_UPLOAD','IN_BATCH_FILE','UPLOADED_TO_RAZORPAY')
                """,
            )
            pending_batch_uploads = await c.fetchval(
                """
                SELECT COUNT(*) FROM rzp_kyc_batches
                 WHERE status IN ('GENERATED','DOWNLOADED') AND record_count > 0
                """,
            )

        submissions_by_status = {r["status"]: int(r["n"]) for r in sub_counts}
        batches_by_status     = {r["status"]: int(r["n"]) for r in batch_counts}

        total_submissions = sum(submissions_by_status.values())

        oldest_age_hours: Optional[float] = None
        if oldest_pending and oldest_pending["slot_at"]:
            age = datetime.now(timezone.utc) - oldest_pending["slot_at"]
            oldest_age_hours = round(age.total_seconds() / 3600, 2)

        return {
            "submissions": {
                "total":                  total_submissions,
                "pending_upload":         submissions_by_status.get("PENDING_BATCH_UPLOAD", 0),
                "in_batch_file":          submissions_by_status.get("IN_BATCH_FILE", 0),
                "uploaded_to_razorpay":   submissions_by_status.get("UPLOADED_TO_RAZORPAY", 0),
                "approved":               submissions_by_status.get("APPROVED", 0),
                "rejected":               submissions_by_status.get("REJECTED", 0),
            },
            "batches":              batches_by_status,
            "pending_batch_uploads": int(pending_batch_uploads or 0),
            "pending_merchants":     int(pending_merchants or 0),
            "oldest_pending_batch": (
                {
                    "batch_id":     oldest_pending["id"],
                    "batch_no":     oldest_pending["batch_no"],
                    "slot_at":      oldest_pending["slot_at"].isoformat(),
                    "record_count": oldest_pending["record_count"],
                    "age_hours":    oldest_age_hours,
                }
                if oldest_pending else None
            ),
            "alerts": _build_alerts(oldest_age_hours, pending_batch_uploads, pending_merchants),
        }


# ── alerts ──────────────────────────────────────────────────────────────────
def _build_alerts(
    oldest_age_hours: Optional[float],
    pending_batches: Optional[int],
    pending_merchants: Optional[int],
) -> list[dict]:
    alerts: list[dict] = []
    if not oldest_age_hours:
        return alerts
    if oldest_age_hours >= 8:
        alerts.append({
            "level": "CRITICAL",
            "code":  "MERCHANT_ONBOARDING_SLA_BREACHED",
            "message": (
                f"Oldest pending batch is {oldest_age_hours}h old — "
                f"{pending_merchants} merchant(s) past the 4-8h SLA window."
            ),
        })
    elif oldest_age_hours >= 2:
        alerts.append({
            "level": "HIGH",
            "code":  "BATCH_UPLOAD_DELAYED",
            "message": (
                f"Batch unprocessed for {oldest_age_hours}h. "
                f"{pending_merchants} merchant onboarding(s) delayed."
            ),
        })
    elif oldest_age_hours >= 0.5:
        alerts.append({
            "level": "WARN",
            "code":  "BATCH_PENDING_UPLOAD",
            "message": (
                f"{pending_batches} batch(es) awaiting Razorpay Dashboard upload."
            ),
        })
    return alerts


# ── CSV / XLSX rendering ────────────────────────────────────────────────────
def _csv_value(row: dict, col: str) -> Any:
    v = row.get(col)
    if v is None:
        return ""
    # Razorpay expects 0/1 ints, not bools.
    if col in {"dashboard_access", "customer_refunds"}:
        return int(v)
    return v


def _render_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow(CSV_COLUMNS)
    for r in rows:
        writer.writerow([_csv_value(r, c) for c in CSV_COLUMNS])
    return buf.getvalue().encode("utf-8")


def _render_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "linked_accounts"
    ws.append(CSV_COLUMNS)
    # account_number must be written as text so Excel doesn't render long
    # values like 918010036353168 as "9.1801E+14" and doesn't strip any
    # leading zeros. ifsc_code is also forced text for safety.
    text_cols = {"account_number", "ifsc_code"}
    for r in rows:
        values = []
        for c in CSV_COLUMNS:
            v = _csv_value(r, c)
            if c in text_cols and v != "":
                v = str(v)
            values.append(v)
        ws.append(values)
        last_row = ws.max_row
        for c in text_cols:
            col_idx = CSV_COLUMNS.index(c) + 1
            ws.cell(row=last_row, column=col_idx).number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Singleton
rzp_kyc_batch_service = RzpKycBatchService()
